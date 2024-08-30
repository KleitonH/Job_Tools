import pyautogui # Importação da biblioteca Pyautogui para controle do mouse e teclado 
import time # Importação da biblioteca time para intervalos de tempo 
import re
import openpyxl
import Breaker
from pytesseract import pytesseract # Importação da biblioteca pytesseract para conversão de imagens em strings
caminho_tesseract = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe" # Caminho para o executável do pytesseract
pytesseract.tesseract_cmd = caminho_tesseract # Define o caminho para o pytesseract
caminho_arquivo = "C:\\VendasPerdidasRepetidas.xlsx"
workbook = openpyxl.load_workbook(caminho_arquivo)
sheet = workbook.active

tempo_de_execucao = 40000
contador_linhas = 14
y_linha = 148
lista_verificacao = []
lista_repetidos = []
codigo_anterior = ""
repeticoes_codigo_anterior = 0
#FILTROS para cada tipo de célula
filtro_codigo_interno = r'[a-zA-Z@,!#./]'
filtro_codigo_fabricante = r'[@!#.,°/]'
filtro_data = r'[a-zA-Z@,!#.]' 
filtro_marca = r'[@,!#./|:;]' 
filtro_classe = r'[a-zG-Z0-9@,!#./%|]' 
filtro_descricao = r'[@!#%]' 


def screenshotCelula(nome_arquivo, filtro, X, Largura, Altura):
    screenshot = pyautogui.screenshot(region=(X , y_linha, Largura, Altura)) # Variável para receber código da célula do item da tabela
    screenshot = screenshot.convert("L")  # Converte para escala de cinza
    screenshot.save(f'./screenshots/screenshots_lostsale/{nome_arquivo}.png') # Salva a imagem

    from PIL import Image # Interpretação de imagem
    screenshot = Image.open(f'./screenshots/screenshots_lostsale/{nome_arquivo}.png') # Recebimento da imagem
    textcod = pytesseract.image_to_string(screenshot) # Converte para texto
    valor_variavel = re.sub(filtro, '', textcod) #Limpa qualquer algarismo indesejado interpretado pelo Tesseract
    print (valor_variavel)
    return valor_variavel

print("Bem-vindo ao SectionDescripted, o programa que classifica grupos e subgrupos com base nas descrições de itens.") # Mensagem de boas vindas
option = "" # Variável que receberá a opção selecionada
while option != "2":
    print("Para começar, analise a opção desejada: \n 1 - Iniciar classificação \n 2 - Sair do programa") # Mensagem de seleção de opções
    option = input("Digite o número da opção desejada: ") 
    if option == "1":
        Breaker.reiniciar_monitoramento()
        print("_____________________________________________________________________________")
        print("Iniciando operação de classificação de itens em 5 segundos...") 
        time.sleep(1) #
        print("4 segundos...")
        time.sleep(1)
        print("3 segundos...")
        time.sleep(1)
        print("2 segundos...")
        time.sleep(1)
        print("1 segundo...")
        time.sleep(1)
        print("Iniciando operação...") 
        tempo_inicial = time.time() # Regra para definir o período de tempo
        contadorrepeticoes = 0
        monitor_thread = Breaker.iniciar_monitoramento()
        while(time.time() - tempo_inicial) < tempo_de_execucao and repeticoes_codigo_anterior < 6 and Breaker.verificar_interrupcao():
            adicionar_tupla = False
            time.sleep(0.3)
            codigo_interno = screenshotCelula("codigointerno", filtro_codigo_interno, 99, 83, 17)
            if codigo_interno != "":
                if codigo_interno in lista_repetidos:
                    pass
                if codigo_interno in lista_verificacao:
                    lista_verificacao.remove(codigo_interno)
                    lista_repetidos.append(codigo_interno)
                    adicionar_tupla = True  # Sinaliza que devemos adicionar a tupla
                else:
                    lista_verificacao.append(codigo_interno)

            if adicionar_tupla:
                data_venda = screenshotCelula("datavenda", filtro_data, 30, 68, 17)
                codigo_fabricante = screenshotCelula("codigofabricante", filtro_codigo_fabricante, 184, 100, 17)
                marca = screenshotCelula("marcaitem", filtro_marca, 287, 55, 17)
                classe = screenshotCelula("classe", filtro_classe, 344, 40, 17)
                descricao = screenshotCelula("descricaoitem", filtro_descricao, 389, 375, 17)

                proxima_linha = sheet.max_row + 1
                
                sheet.cell(row=proxima_linha, column=1).value = data_venda
                sheet.cell(row=proxima_linha, column=2).value = codigo_fabricante
                sheet.cell(row=proxima_linha, column=3).value = marca
                sheet.cell(row=proxima_linha, column=4).value = classe
                sheet.cell(row=proxima_linha, column=5).value = descricao
                
                workbook.save(caminho_arquivo)
                
                print(f"Dados adicionados na linha {proxima_linha} da planilha.")
                contadorrepeticoes += 1

            else:
                pass
            
            if contador_linhas > 0:
                y_linha += 16

            if codigo_interno == codigo_anterior or codigo_anterior == "":
                repeticoes_codigo_anterior +=1
            elif codigo_interno != codigo_anterior:
                repeticoes_codigo_anterior = 0

            if contador_linhas > 0:
                contador_linhas -= 1
            codigo_anterior = codigo_interno
            time.sleep(1.0)
            pyautogui.press('down')
        exit()

    elif option == "2": # Se a opção for 4, sai do programa
        print("Saindo do programa...")
        exit()


            
            

