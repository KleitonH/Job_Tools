import pyautogui
import pyperclip
import openpyxl
import time
import Breaker

# Caminho do arquivo Excel
caminho_arquivo = r"C:\Users\Usuario\Desktop\cadastros.xlsx"
wb = openpyxl.load_workbook(caminho_arquivo)
ws = wb.active

# Função para clicar duas vezes, copiar e obter o texto
def capturar_texto(x, y):
    try:
        pyautogui.tripleClick(x, y)  # Clica duas vezes no campo
        time.sleep(0.3)  # Espera um pouco para garantir a seleção
        pyautogui.hotkey('ctrl', 'c')  # Copia o conteúdo
        time.sleep(0.3)
        return pyperclip.paste().strip()  # Retorna o texto copiado
    except Exception as e:
        print(f"Erro ao copiar texto: {e}")
        return ""

# Começa a partir da linha 115
linha = 124
monitor_thread = Breaker.iniciar_monitoramento()
time.sleep(3)
while Breaker.verificar_interrupcao():
    cnpj = ws[f"H{linha}"].value
    if not cnpj:
        break  # Se não houver CNPJ, encerra

    # Automação de inserção do CNPJ
    pyautogui.press('f2')  # Ativa edição
    pyautogui.write(str(cnpj))
    pyautogui.press('enter', presses=2)
    time.sleep(1)

    # Capturar e inserir os valores nas células
    ws[f"J{linha}"] = capturar_texto(790, 485)  # Posição central da área (697, 478, 179, 17)
    ws[f"K{linha}"] = capturar_texto(717, 510)  # (695, 502, 44, 15)
    ws[f"L{linha}"] = capturar_texto(790, 557)  # (697, 551, 184, 13)
    ws[f"M{linha}"] = capturar_texto(788, 582)  # (695, 575, 186, 14)
    ws[f"N{linha}"] = capturar_texto(790, 388)  # (697, 381, 183, 14)
    time.sleep(1.2)

    # Fechar interface do Tesseract (se necessário)
    pyautogui.press('esc')
    pyautogui.press('enter')
    pyautogui.press('esc')

    # Avançar para a próxima linha
    linha += 1

# Salvar a planilha
wb.save(caminho_arquivo)
print("Processo concluído!")